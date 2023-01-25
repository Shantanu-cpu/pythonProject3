import openpyxl

def get_no_rows(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_no_columns(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


def read_data(file_name, sheet_name, rownum, colnum):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(rownum, colnum).value


def write_data(file_name, sheet_name, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(rownum, colnum).value = data
    workbook.save(file_name)